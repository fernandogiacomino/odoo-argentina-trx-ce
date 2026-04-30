# Part of l10n-ar-edi-community. See LICENSE file for full copyright and licensing details.
"""Writer XLSX para Subdiario IVA Compras/Ventas.

Lib pura — no importa ``odoo``. Recibe estructuras Python (dicts y listas)
y devuelve ``bytes`` con el XLSX listo para servir. Wrapper finito sobre
``openpyxl`` para encapsular el lookup lazy del import (así un Odoo
sin openpyxl no rompe a la importación del módulo, sino que da un mensaje
claro al ejecutar el wizard).

Estilo de salida (decidido con Hector 2026-04-26):
    - Header en negrita + fill gris suave + border bottom doble.
    - Importes con format ``#,##0.00`` y alineados a la derecha.
    - Fila de totales con border top, fondo amarillo claro, negrita.
    - Auto-width pragmático: ancho = max(longitud, header_len) clamp 8..50.
    - Una hoja por sección: ``Ventas`` y ``Compras`` (si hay datos).
    - Hoja ``Resumen`` opcional con totales por concepto.

Las columnas son configurables: se pasa una lista de spec dicts:
    {"key": "fecha", "header": "Fecha", "width": 12, "type": "date"}
    {"key": "neto_gravado", "header": "Neto Gravado", "type": "amount"}
    {"key": "tipo_cbte", "header": "Tipo", "type": "text"}

Tipos soportados: ``text``, ``int``, ``amount`` (float, format ARS),
``percent`` (float, format %), ``date`` (datetime.date).
"""
import io
import logging

_logger = logging.getLogger(__name__)


# Tipos de columna que sumamos en la fila de totales.
_NUMERIC_TYPES = {"amount", "int", "percent"}


def _require_openpyxl():
    """Lazy import — si falta, da un error humano antes que un ImportError feo.

    Se separa para que (a) testear sin openpyxl no rompa colección, (b) el
    error sea accionable cuando el wizard se llama en un Odoo que no lo tiene.
    """
    try:
        import openpyxl  # noqa: F401
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "El módulo Python 'openpyxl' no está instalado en este Odoo. "
            "Instalalo con 'pip install openpyxl' (o agregalo al "
            "requirements.txt del container)."
        ) from exc
    return openpyxl, Alignment, Border, Font, PatternFill, Side, get_column_letter


def build_xlsx(sheets, meta=None):
    """Construye un XLSX en memoria y devuelve bytes.

    :param sheets: lista de dicts por hoja:
        {
            "name": "Ventas",
            "title": "Subdiario IVA Ventas — Trixocom — 2026-04",
            "columns": [
                {"key": "fecha", "header": "Fecha", "type": "date", "width": 12},
                {"key": "tipo_cbte", "header": "Tipo", "type": "text", "width": 8},
                {"key": "neto_gravado", "header": "Neto Gravado", "type": "amount"},
                ...
            ],
            "rows": [{"fecha": date(...), "tipo_cbte": "FA-A", ...}, ...],
            "totals": True,    # default True — agrega fila Totales
        }
    :param meta: dict opcional con metadatos {"author": str, "title": str}.
    :return: bytes del archivo .xlsx
    """
    openpyxl, Alignment, Border, Font, PatternFill, Side, col_letter = _require_openpyxl()

    wb = openpyxl.Workbook()
    # Eliminar la hoja por defecto que viene vacía.
    default = wb.active
    wb.remove(default)

    if meta:
        wb.properties.title = meta.get("title", "")
        wb.properties.creator = meta.get("author", "Odoo / Trixocom")

    if not sheets:
        # XLSX vacío: dejamos una hoja "Sin datos" para que el archivo
        # abra sin error en Excel/LibreOffice.
        ws = wb.create_sheet("Sin datos")
        ws["A1"] = "No hay comprobantes en el período seleccionado."
        return _wb_to_bytes(wb)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="334E68")  # azul oscuro
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(bold=True, size=12)
    totals_font = Font(bold=True)
    totals_fill = PatternFill("solid", fgColor="FFF8C5")  # amarillo claro
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    border_top = Border(top=medium)
    border_header = Border(bottom=thin)

    for sheet_spec in sheets:
        ws = wb.create_sheet(sheet_spec["name"][:31])  # excel limita 31 chars
        cols = sheet_spec["columns"]
        rows = sheet_spec.get("rows", [])
        title = sheet_spec.get("title")

        row_cursor = 1

        # Título arriba (merged).
        if title:
            ws.cell(row=row_cursor, column=1, value=title).font = title_font
            ws.merge_cells(
                start_row=row_cursor, start_column=1,
                end_row=row_cursor, end_column=len(cols),
            )
            ws.row_dimensions[row_cursor].height = 22
            row_cursor += 2  # fila en blanco entre título y headers

        header_row = row_cursor
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=header_row, column=ci, value=col["header"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border_header
            # Ancho.
            width = col.get("width") or _infer_width(col)
            ws.column_dimensions[col_letter(ci)].width = width
        ws.row_dimensions[header_row].height = 30

        # Filas de datos.
        data_start = header_row + 1
        for ri, data in enumerate(rows, start=data_start):
            for ci, col in enumerate(cols, start=1):
                value = data.get(col["key"])
                cell = ws.cell(row=ri, column=ci)
                _write_cell(cell, value, col, Alignment)

        # Totales.
        if rows and sheet_spec.get("totals", True):
            tr = data_start + len(rows)
            for ci, col in enumerate(cols, start=1):
                cell = ws.cell(row=tr, column=ci)
                cell.font = totals_font
                cell.fill = totals_fill
                cell.border = border_top
                if ci == 1:
                    cell.value = "TOTALES"
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col.get("type") in _NUMERIC_TYPES and col.get("type") != "percent":
                    # =SUM(C2:C10) — referencia a la columna data.
                    letter = col_letter(ci)
                    cell.value = f"=SUM({letter}{data_start}:{letter}{tr - 1})"
                    cell.number_format = _number_format_for(col)
                    cell.alignment = Alignment(horizontal="right")

        # Freeze: header + título.
        ws.freeze_panes = ws.cell(row=data_start, column=1)

    return _wb_to_bytes(wb)


def _wb_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _infer_width(col):
    """Ancho aproximado en chars Excel."""
    t = col.get("type", "text")
    base = {
        "date": 11,
        "int": 8,
        "amount": 14,
        "percent": 8,
        "text": 18,
    }.get(t, 14)
    header_len = len(col.get("header") or "")
    return min(50, max(8, max(base, header_len + 2)))


def _number_format_for(col):
    t = col.get("type", "text")
    if t == "amount":
        # Formato ARS sin símbolo, separador de miles + 2 decimales,
        # negativos en rojo entre paréntesis.
        return '#,##0.00;[Red](#,##0.00)'
    if t == "percent":
        # 0.21 → 21.00%. Pero como nuestros datos vienen en pct ya (21.0),
        # usamos "0.00%" sin el factor (multiplicar por 1).
        # Para no liar, dejamos numerico simple.
        return '0.00'
    if t == "int":
        return '0'
    return None


def _write_cell(cell, value, col, Alignment):
    t = col.get("type", "text")
    if value is None:
        cell.value = None
        return
    if t == "amount":
        try:
            cell.value = float(value)
        except (TypeError, ValueError):
            cell.value = 0.0
        cell.number_format = _number_format_for(col)
        cell.alignment = Alignment(horizontal="right")
    elif t == "int":
        try:
            cell.value = int(value)
        except (TypeError, ValueError):
            cell.value = 0
        cell.number_format = _number_format_for(col)
        cell.alignment = Alignment(horizontal="right")
    elif t == "percent":
        try:
            cell.value = float(value)
        except (TypeError, ValueError):
            cell.value = 0.0
        cell.number_format = _number_format_for(col)
        cell.alignment = Alignment(horizontal="right")
    elif t == "date":
        cell.value = value
        cell.number_format = "yyyy-mm-dd"
        cell.alignment = Alignment(horizontal="center")
    else:  # text
        cell.value = str(value)
        cell.alignment = Alignment(horizontal="left", wrap_text=False)
